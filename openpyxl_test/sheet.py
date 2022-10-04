from openpyxl import Workbook, load_workbook
from io import BytesIO

# Criando um arquivo de planilha

wb = Workbook()
ws = wb.active

dados = [
    ('Nome', 'Idade', 'Sexo'),
    ('Eric', '22', 'M'),
    ('Guilherme', '22', 'M'),
    ('Leonardo', '22', 'M'),
]

for dado in dados:
    ws.append(dado)

wb.save('sheet.xlsx')

# Escrevendo planilha em stream de bytes

bytestream = BytesIO()
wb.save(bytestream)
bytestream.seek(0)
with open('bytes_sheet.xlsx', 'wb') as f:
    f.write(bytestream.read())

# Lendo arquivo de planilha

wb = load_workbook('sheet.xlsx')
ws = wb.active

for row in ws.values:
    print(row)

dados = []
values = list(ws.values)
headers = values[0]
for row in values:
    dados.append({col: row[i] for i, col in enumerate(headers)})

for row in dados:
    print(row)